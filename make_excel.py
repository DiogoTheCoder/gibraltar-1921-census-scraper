#!/usr/bin/env python3
"""Convert the scraped CSV into a formatted .xlsx (frozen header + auto-filter).

Usage:
    pip install openpyxl
    python3 make_excel.py [input.csv] [output.xlsx]
"""
import csv
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

WIDTHS = {
    "ID": 7, "Division": 8, "District": 8, "HouseNo": 8, "Address": 20,
    "NoOfPersons": 11, "Surname": 16, "Forename": 14, "RelationToHead": 15,
    "MaritalStatus": 13, "Sex": 5, "Age": 5, "Occupation": 22, "Employer": 12,
    "Worker": 10, "WorkingOwnAccount": 18, "Birthplace": 18, "Religion": 16,
    "Education": 18, "Disabilities": 12, "No_of_Rooms": 11, "Remarks": 14,
}


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "gibraltar_1921_census.csv"
    dst = sys.argv[2] if len(sys.argv) > 2 else "gibraltar_1921_census.xlsx"

    rows = list(csv.DictReader(open(src, encoding="utf-8")))
    fields = list(rows[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "1921 Census"
    ws.append(fields)

    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="305496")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")

    for r in rows:
        ws.append([r[k] for k in fields])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows) + 1}"
    for i, k in enumerate(fields, 1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(k, 14)

    wb.save(dst)
    print(f"Wrote {len(rows)} rows to {dst}")


if __name__ == "__main__":
    main()
