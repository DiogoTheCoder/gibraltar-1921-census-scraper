#!/usr/bin/env python3
"""Build the 1921 Gibraltar Census workbook: Summary -> Geography -> Raw Data.

- Summary sheet (first): headline totals plus breakdowns by Ward, Sex, Religion,
  Marital Status, Education, and top Birthplaces / Occupations.
- Geography sheet: Ward -> Division -> District -> headcount + representative
  streets. Wards/police districts come from the City Council Ordinance 1921,
  First Schedule (List of Wards).
- Raw Data sheet (last): every scraped record, EXACTLY as in the CSV (no added
  columns).

Usage:
    pip install openpyxl
    python3 build_workbook.py [input.csv] [output.xlsx]
"""
import collections
import csv
import statistics
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Census "Division" == Ward (Divisions 5 & 6 are special, non-ward enumerations).
WARD = {
    "1": "Old Town Ward",
    "2": "Castle Ward",
    "3": "Cathedral Ward",
    "4": "Europa Ward",
    "5": "Military & Government establishments",
    "6": "Shipping (afloat)",
}
WARD_ORDER = ["1", "2", "3", "4", "5", "6"]

# --- Franchise estimate (City Council Ordinance 1921, ss. 14-15) -------------
# Electors = British subject, of full age (21+), an occupier of premises as
# owner/tenant/lodger; women excluded; Crown servants in rent-free quarters
# excluded. The census records none of nationality, rates, legal incapacity or
# tenure directly, so these are proxies (see the notes written onto the sheet).
ELECTOR_RELATIONS = {"Head", "Head of family", "Lodger"}
ELECTOR_DIVISIONS = ["1", "2", "3", "4"]  # wards; Div 5 (Crown quarters) & 6 (afloat) excluded
FULL_AGE = 21

_BRITISH_PLACES = {
    "GIBRALTAR", "GIBRLATAR", "GIBRALTA", "GIB", "ENGLAND", "ENG", "SCOTLAND",
    "IRELAND", "WALES", "BRITAIN", "UK", "MALTA", "GOZO", "INDIA", "CYPRUS",
    "JERSEY", "GUERNSEY", "CANADA", "AUSTRALIA", "JAMAICA", "CEYLON",
    "SINGAPORE", "ADEN", "HONGKONG",
}
_FOREIGN_PLACES = {
    "SPAIN", "PORTUGAL", "ITALY", "FRANCE", "MOROCCO", "MAROC", "TETUAN",
    "MARAKESH", "MOGADOR", "JAPAN", "GERMANY", "AUSTRIA", "GREECE", "USA",
    "AMERICA", "ARGENTINA", "TURKEY", "RUSSIA", "HOLLAND", "BELGIUM", "SWEDEN",
    "NORWAY", "DENMARK", "SWITZERLAND", "BRAZIL", "CHINA", "EGYPT", "POLAND",
    "CUBA", "MEXICO", "CHILE", "URUGUAY", "PHILIPPINES",
}


def british_subject(birthplace):
    """Classify birthplace as 'british' / 'foreign' / 'uncertain'.

    The census annotates foreign-born British subjects with a 'BS' suffix
    (e.g. 'La Linea SPAIN BS') and non-subjects with 'NBS'.
    """
    t = birthplace.strip().upper().split()
    if not t or t == ["NO", "DATA"]:
        return "uncertain"
    if t[-1] == "BS" or ("BRITISH" in t and "SUBJECT" in t):
        return "british"
    if "NBS" in t or "(NBS)" in t or ("NOT" in t and "SUBJECT" in t):
        return "foreign"
    if any(w in _BRITISH_PLACES for w in t):
        return "british"
    if any(w in _FOREIGN_PLACES for w in t):
        return "foreign"
    return "uncertain"


TITLE_FONT = Font(bold=True, size=14)
H1_FONT = Font(bold=True, size=12, color="305496")
HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="305496")
SUBTLE = Font(italic=True, color="808080")


def style_header_row(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center")


def pct(n, total):
    return round(100.0 * n / total, 1) if total else 0.0


def add_table(ws, start_row, headers, rows, widths=None):
    """Write a header + rows starting at start_row; return next free row."""
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for j, v in enumerate(row, 1):
            ws.cell(row=r, column=j, value=v)
        r += 1
    return r + 1  # blank spacer line


def counter_rows(rows, key, total, top=None, drop_nodata=False):
    c = collections.Counter(r[key].strip() or "(blank)" for r in rows)
    if drop_nodata:
        c.pop("No Data", None)
        c.pop("(blank)", None)
    items = c.most_common(top) if top else sorted(c.items(),
                                                  key=lambda kv: -kv[1])
    return [[k, v, pct(v, total)] for k, v in items]


def build_summary(ws, rows):
    total = len(rows)
    ws["A1"] = "1921 Gibraltar Census — Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Source: Gibraltar National Archives (nationalarchives.gi). "
                "Wards per City Council Ordinance 1921, First Schedule.")
    ws["A2"].font = SUBTLE

    males = sum(1 for r in rows if r["Sex"] == "M")
    females = sum(1 for r in rows if r["Sex"] == "F")
    ages = [int(r["Age"]) for r in rows if r["Age"].isdigit()]
    surnames = len({r["Surname"] for r in rows})

    r = add_table(ws, 4, ["Metric", "Value"], [
        ["Total individuals", total],
        ["Distinct surnames", surnames],
        ["Male", f"{males} ({pct(males, total)}%)"],
        ["Female", f"{females} ({pct(females, total)}%)"],
        ["Age — min / max", f"{min(ages)} / {max(ages)}"],
        ["Age — mean", round(statistics.mean(ages), 1)],
        ["Age — median", int(statistics.median(ages))],
        ["Records with non-numeric age", total - len(ages)],
    ])

    # By Ward (Division)
    ws.cell(row=r, column=1, value="Population by Ward").font = H1_FONT
    r += 1
    ward_rows = []
    for dv in WARD_ORDER:
        sub = [x for x in rows if x["Division"] == dv]
        if not sub:
            continue
        m = sum(1 for x in sub if x["Sex"] == "M")
        f = sum(1 for x in sub if x["Sex"] == "F")
        ward_rows.append([WARD[dv], dv, len(sub), pct(len(sub), total), m, f])
    r = add_table(ws, r, ["Ward", "Division", "People", "% of total",
                          "Male", "Female"], ward_rows)

    # Simple categorical breakdowns
    for title, key, top, drop in [
        ("By Sex", "Sex", None, False),
        ("By Religion", "Religion", None, False),
        ("By Marital Status", "MaritalStatus", None, False),
        ("By Education", "Education", None, False),
        ("Top 15 Birthplaces", "Birthplace", 15, False),
        ("Top 20 Occupations (excl. 'No Data')", "Occupation", 20, True),
    ]:
        ws.cell(row=r, column=1, value=title).font = H1_FONT
        r += 1
        r = add_table(ws, r, [key, "People", "% of total"],
                      counter_rows(rows, key, total, top=top, drop_nodata=drop))

    for col, w in {"A": 42, "B": 14, "C": 12, "D": 12, "E": 10, "F": 10}.items():
        ws.column_dimensions[col].width = w


def build_electors(ws, rows):
    ws["A1"] = "Estimated Electors per Ward — 1921 Franchise"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("City Council Ordinance 1921, ss. 14-15. ESTIMATE from census "
                "fields (proxies) — see notes below.")
    ws["A2"].font = SUBTLE

    def qualifies(x, relations):
        return (x["Sex"] == "M"
                and x["Age"].isdigit() and int(x["Age"]) >= FULL_AGE
                and x["Division"] in ELECTOR_DIVISIONS
                and x["RelationToHead"].strip() in relations)

    def ward_table(relations):
        out = []
        te = to = tu = 0
        for dv in ELECTOR_DIVISIONS:
            sub = [x for x in rows if x["Division"] == dv and qualifies(x, relations)]
            brit = sum(1 for x in sub if british_subject(x["Birthplace"]) == "british")
            unc = sum(1 for x in sub if british_subject(x["Birthplace"]) == "uncertain")
            out.append([WARD[dv], dv, len(sub), brit, unc])
            te += brit
            to += len(sub)
            tu += unc
        out.append(["TOTAL", "", to, te, tu])
        return out

    hdr = ["Ward", "Division", "Adult male occupiers",
           "Estimated electors (British subject)", "Uncertain nationality"]
    r = add_table(ws, 4, hdr, ward_table(ELECTOR_RELATIONS))

    # per police district within each ward
    ws.cell(row=r, column=1,
            value="By police district (within ward)").font = H1_FONT
    r += 1
    dhdr = ["Ward", "Division", "District", "Adult male occupiers",
            "Estimated electors (British subject)", "Uncertain nationality"]
    drows = []
    for dv in ELECTOR_DIVISIONS:
        dists = sorted({x["District"] for x in rows if x["Division"] == dv},
                       key=lambda d: (not d.isdigit(), int(d) if d.isdigit() else 0))
        for dist in dists:
            sub = [x for x in rows if x["Division"] == dv and x["District"] == dist
                   and qualifies(x, ELECTOR_RELATIONS)]
            if not sub:
                continue
            brit = sum(1 for x in sub if british_subject(x["Birthplace"]) == "british")
            unc = sum(1 for x in sub if british_subject(x["Birthplace"]) == "uncertain")
            drows.append([WARD[dv], dv, dist, len(sub), brit, unc])
    r = add_table(ws, r, dhdr, drows)

    # sensitivity line
    ws.cell(row=r, column=1,
            value="Sensitivity — if 'Boarder' is also treated as a lodger:"
            ).font = H1_FONT
    r += 1
    r = add_table(ws, r, hdr,
                  ward_table(ELECTOR_RELATIONS | {"Boarder"}))

    notes = [
        "How this estimate maps the franchise onto census fields:",
        "  • British subject  -> Birthplace in a British territory, or annotated 'BS' "
        "(the census marks foreign-born subjects 'BS' and non-subjects 'NBS').",
        "  • Of full age      -> Age >= 21.",
        "  • Women excluded    -> Sex = M only (a franchise proviso).",
        "  • Owner/tenant/lodger -> RelationToHead is Head, Head of family or Lodger "
        "(household members who are not the occupier do not qualify in their own right).",
        "  • Crown rent-free quarters excluded -> Divisions 5 (military & government "
        "establishments) and 6 (afloat) are dropped entirely.",
        "",
        "Cannot be tested from the census (assumed satisfied, so this is an UPPER bound):",
        "  • Not in arrears of rates;  • Not under legal incapacity;  • >= 6 months' occupation.",
        "'Uncertain nationality' = birthplace missing or unclassifiable; electors likely lie "
        "between the British-subject figure and that figure plus the uncertain count.",
    ]
    r += 1
    for line in notes:
        c = ws.cell(row=r, column=1, value=line)
        if line.endswith(":"):
            c.font = Font(bold=True)
        r += 1

    for col, w in {"A": 40, "B": 10, "C": 10, "D": 22,
                   "E": 34, "F": 22}.items():
        ws.column_dimensions[col].width = w


def build_geography(ws, rows):
    ws["A1"] = "Geography — Ward / Division / Police District"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Police districts per City Council Ordinance 1921. Numbers "
                "restart in Europa Ward (the '(South)' series), so read "
                "Division + District together.")
    ws["A2"].font = SUBTLE

    # group by (division, district)
    groups = collections.defaultdict(list)
    for x in rows:
        groups[(x["Division"], x["District"])].append(x)

    def dkey(item):
        (dv, dist) = item
        return (dv, not dist.isdigit(), int(dist) if dist.isdigit() else 0)

    table = []
    for (dv, dist) in sorted(groups, key=dkey):
        sub = groups[(dv, dist)]
        streets = collections.Counter(x["Address"].strip() for x in sub)
        top = " · ".join(f"{s} ({n})" for s, n in streets.most_common(6))
        table.append([WARD.get(dv, dv), dv, dist, len(sub), top])

    r = add_table(ws, 4, ["Ward", "Division", "District", "People",
                          "Representative streets (headcount)"], table)
    for col, w in {"A": 34, "B": 10, "C": 10, "D": 10, "E": 90}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


def build_raw(ws, rows, fields):
    ws.append(fields)
    style_header_row(ws, 1, len(fields))
    for r in rows:
        ws.append([r[k] for k in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows) + 1}"
    widths = {"ID": 7, "Division": 8, "District": 8, "HouseNo": 8, "Address": 20,
              "NoOfPersons": 11, "Surname": 16, "Forename": 14,
              "RelationToHead": 15, "MaritalStatus": 13, "Sex": 5, "Age": 5,
              "Occupation": 22, "Employer": 12, "Worker": 10,
              "WorkingOwnAccount": 18, "Birthplace": 18, "Religion": 16,
              "Education": 18, "Disabilities": 12, "No_of_Rooms": 11,
              "Remarks": 14}
    for i, k in enumerate(fields, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(k, 14)


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "gibraltar_1921_census.csv"
    dst = sys.argv[2] if len(sys.argv) > 2 else "gibraltar_1921_census.xlsx"
    rows = list(csv.DictReader(open(src, encoding="utf-8")))
    fields = list(rows[0].keys())

    wb = Workbook()
    build_summary(wb.active, rows)
    wb.active.title = "Summary"
    build_electors(wb.create_sheet("Electors"), rows)
    build_geography(wb.create_sheet("Geography"), rows)
    build_raw(wb.create_sheet("Raw Data"), rows, fields)
    wb.save(dst)
    print(f"Wrote {dst}: Summary + Electors + Geography + Raw Data "
          f"({len(rows)} records)")


if __name__ == "__main__":
    main()
